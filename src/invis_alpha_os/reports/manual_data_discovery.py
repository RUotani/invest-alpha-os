"""Discover local manual/broker JP bars data files (CSV/TSV/TXT/XLSX, paths redacted)."""

from __future__ import annotations

import fnmatch
import hashlib
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from invis_alpha_os.config.env_file_loader import is_git_tracked
from invis_alpha_os.reports.manual_csv_discovery import _location_label, _search_roots
from invis_alpha_os.reports.manual_csv_pii_guard import run_manual_csv_pii_guard, scan_csv_headers_for_pii
from invis_alpha_os.reports.manual_data_dropzone import is_excluded_manual_filename
from invis_alpha_os.reports.manual_data_recent_candidates import scan_recent_ohlcv_candidates
from invis_alpha_os.reports.manual_data_schema_probe import probe_path_ohlcv_schema

SUPPORTED_EXTENSIONS: tuple[str, ...] = (".csv", ".tsv", ".txt", ".xlsx")

EXACT_CANDIDATE_NAMES: tuple[str, ...] = (
    "manual_jp_bars.csv",
    "manual_jp_bars.tsv",
    "manual_jp_bars.xlsx",
    "manual_jp_bars.txt",
    "jp_bars_manual.csv",
    "jp_bars.csv",
    "jp_bars.tsv",
    "jp_daily_bars.csv",
    "broker_jp_bars.csv",
    "moomoo_jp_bars.csv",
    "sbi_jp_bars.csv",
    "rakuten_jp_bars.csv",
    "kabu_jp_bars.csv",
)

GLOB_NAME_PATTERNS: tuple[str, ...] = (
    "jp_bars.*",
    "jp_daily_bars.*",
    "broker_jp_bars.*",
    "moomoo_jp_bars.*",
    "sbi_jp_bars.*",
    "rakuten_jp_bars.*",
    "kabu_jp_bars.*",
    "kabuka.*",
    "株価.*",
)


def _openpyxl_available() -> bool:
    try:
        import importlib.util

        return importlib.util.find_spec("openpyxl") is not None
    except ImportError:
        return False


def _read_xlsx_headers(path: Path) -> tuple[list[str], list[str]]:
    if not _openpyxl_available():
        return [], ["xlsx_not_supported_no_openpyxl"]
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        workbook.close()
        if row is None:
            return [], ["xlsx_empty"]
        return [str(cell).strip() for cell in row if cell is not None and str(cell).strip()], []
    except Exception as exc:
        return [], [f"xlsx_read_failed:{exc.__class__.__name__}"]


def _pii_guard_for_data_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        headers, errors = _read_xlsx_headers(path)
        if errors:
            return {
                "pii_guard_status": "skipped",
                "account_data_detected": False,
                "warnings": errors,
            }
        result = scan_csv_headers_for_pii(headers)
        return result.json_payload
    if suffix in {".csv", ".tsv", ".txt"}:
        result = run_manual_csv_pii_guard(path)
        return result.json_payload
    return {"pii_guard_status": "rejected", "account_data_detected": False, "warnings": ["unsupported_extension"]}


def _matches_candidate_name(name: str) -> bool:
    lowered = name.lower()
    if name in EXACT_CANDIDATE_NAMES or lowered in {n.lower() for n in EXACT_CANDIDATE_NAMES}:
        return True
    return any(fnmatch.fnmatch(lowered, pattern.lower()) for pattern in GLOB_NAME_PATTERNS)


def _redacted_header_summary_for_path(path: Path) -> str:
    pii_payload = _pii_guard_for_data_file(path)
    if path.suffix.lower() == ".xlsx":
        headers, _ = _read_xlsx_headers(path)
    elif path.suffix.lower() in {".csv", ".tsv", ".txt"}:
        from invis_alpha_os.reports.manual_csv_pii_guard import read_csv_headers

        headers, _ = read_csv_headers(path)
    else:
        headers = []
    normalized = sorted(h.strip().lower() for h in headers if h and h.strip())
    if not normalized:
        return "cols=0"
    digest = hashlib.sha256(",".join(normalized).encode("utf-8")).hexdigest()[:12]
    return f"cols={len(normalized)};sha256_prefix={digest}"


def _candidate_score(filename: str, *, schema_ohlcv: bool = False) -> int:
    lowered = filename.lower()
    base = 40
    if lowered in {n.lower() for n in EXACT_CANDIDATE_NAMES}:
        base = 100
    elif "manual_jp_bars" in lowered:
        base = 90
    elif "jp_bars" in lowered or "broker_jp_bars" in lowered:
        base = 70
    if schema_ohlcv:
        return max(base, 85)
    return base


def _candidate_record(path: Path, *, repo_root: Path) -> dict[str, Any]:
    pii_payload = _pii_guard_for_data_file(path)
    tracked = is_git_tracked(path, repo_root)
    pii_status = str(pii_payload.get("pii_guard_status", "rejected"))
    account_data = bool(pii_payload.get("account_data_detected"))
    xlsx_blocked = path.suffix.lower() == ".xlsx" and not _openpyxl_available()
    safe = (
        pii_status == "passed"
        and not account_data
        and not tracked
        and path.is_file()
        and not xlsx_blocked
    )
    modified_at: str | None = None
    file_size_bytes: int | None = None
    try:
        stat = path.stat()
        modified_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        file_size_bytes = int(stat.st_size)
    except OSError:
        pass
    safety_status = "pass" if safe else "reject"
    reason = "safe_for_dry_run" if safe else "pii_or_tracked_or_unsupported"
    if account_data:
        reason = "account_data_detected"
    elif tracked:
        reason = "git_tracked_refused"
    elif xlsx_blocked:
        reason = "xlsx_not_supported"
    return {
        "filename": path.name,
        "extension": path.suffix.lower(),
        "path_redacted": True,
        "directory_label": _location_label(path),
        "location_label": _location_label(path),
        "file_size_bytes": file_size_bytes,
        "modified_at": modified_at,
        "candidate_score": _candidate_score(path.name),
        "schema_ohlcv_candidate": False,
        "recent_candidate": False,
        "redacted_header_summary": _redacted_header_summary_for_path(path),
        "safety_status": safety_status,
        "reason": reason,
        "git_tracked": tracked,
        "pii_guard_status": pii_status,
        "account_data_detected": account_data,
        "safe_to_parse": safe,
        "xlsx_supported": _openpyxl_available(),
        "resolved_path": str(path.resolve()),
    }


@dataclass(frozen=True)
class ManualDataDiscoveryResult:
    markdown_text: str
    json_payload: dict[str, Any]
    selected_path: Path | None


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _merge_candidate(
    found: list[dict[str, Any]],
    seen: set[Path],
    path: Path,
    *,
    repo_root: Path,
    schema_ohlcv: bool = False,
    recent: bool = False,
) -> None:
    resolved = path.resolve()
    if resolved in seen or not resolved.is_file():
        return
    if is_excluded_manual_filename(resolved.name):
        return
    seen.add(resolved)
    record = _candidate_record(resolved, repo_root=repo_root)
    record["schema_ohlcv_candidate"] = schema_ohlcv
    record["recent_candidate"] = recent
    record["candidate_score"] = _candidate_score(resolved.name, schema_ohlcv=schema_ohlcv)
    found.append(record)


def discover_manual_data_candidates(
    *,
    repo_root: Path,
    extra_paths: list[Path] | None = None,
    search_roots: list[Path] | None = None,
) -> list[dict[str, Any]]:
    seen: set[Path] = set()
    found: list[dict[str, Any]] = []
    roots = search_roots if search_roots is not None else _search_roots()
    for root in roots:
        if not root.is_dir():
            continue
        for name in EXACT_CANDIDATE_NAMES:
            path = (root / name).resolve()
            if path in seen or not path.is_file():
                continue
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            seen.add(path)
            found.append(_candidate_record(path, repo_root=repo_root))
        try:
            for ext in SUPPORTED_EXTENSIONS:
                for path in sorted(root.glob(f"*{ext}")):
                    resolved = path.resolve()
                    if resolved in seen or not _matches_candidate_name(resolved.name):
                        continue
                    seen.add(resolved)
                    found.append(_candidate_record(resolved, repo_root=repo_root))
        except OSError:
            continue
    for row in scan_recent_ohlcv_candidates(roots):
        _merge_candidate(
            found,
            seen,
            Path(str(row["resolved_path"])),
            repo_root=repo_root,
            schema_ohlcv=bool(row.get("schema_ohlcv_candidate")),
            recent=True,
        )
    for extra in extra_paths or []:
        if extra.is_file():
            schema_ok, _ = probe_path_ohlcv_schema(extra)
            _merge_candidate(found, seen, extra, repo_root=repo_root, schema_ohlcv=schema_ok)
        elif extra.is_dir():
            for name in EXACT_CANDIDATE_NAMES:
                candidate = extra / name
                if candidate.is_file():
                    schema_ok, _ = probe_path_ohlcv_schema(candidate)
                    _merge_candidate(
                        found, seen, candidate, repo_root=repo_root, schema_ohlcv=schema_ok
                    )
    found.sort(
        key=lambda row: (
            -int(row["safe_to_parse"]),
            -int(row.get("schema_ohlcv_candidate", False)),
            -int(row["candidate_score"]),
            row["filename"],
        )
    )
    return found


def _public_payload(candidates: list[dict[str, Any]], selected: dict[str, Any] | None) -> dict[str, Any]:
    public_candidates = [{k: v for k, v in row.items() if k != "resolved_path"} for row in candidates]
    selected_public = None
    if selected is not None:
        selected_public = {k: v for k, v in selected.items() if k != "resolved_path"}
    searched_roots = [root for root in _search_roots() if root.is_dir()]
    has_xlsx_candidate = any(row.get("extension") == ".xlsx" for row in candidates)
    xlsx_supported = _openpyxl_available()
    next_action = (
        "Export OHLCV-only from broker and save as manual_jp_bars.csv in "
        "~/Downloads/invest-alpha-os-manual-data-dropzone"
    )
    if has_xlsx_candidate and not xlsx_supported:
        next_action = "Use CSV/TSV/TXT or install openpyxl locally for XLSX (not bundled in repo)"
    elif selected and selected.get("safe_to_parse"):
        next_action = "Run weekly-candidate-brief-manual-data-import-flow"
    return {
        "candidates_found": len(candidates),
        "searched_location_count": len(searched_roots),
        "supported_extensions": list(SUPPORTED_EXTENSIONS),
        "candidates": public_candidates,
        "selected_candidate": selected_public,
        "safe_to_parse": bool(selected and selected.get("safe_to_parse")),
        "next_required_action": next_action,
        "xlsx_supported": xlsx_supported,
        "contents_printed": False,
        "path_redacted": True,
    }


def build_manual_data_discovery(
    *,
    report_date: str,
    repo_root: Path,
    extra_paths: list[Path] | None = None,
    paste_materialized_path: Path | None = None,
) -> ManualDataDiscoveryResult:
    extras = list(extra_paths or [])
    if paste_materialized_path is not None:
        extras.append(paste_materialized_path)
    candidates = discover_manual_data_candidates(repo_root=repo_root, extra_paths=extras)
    selected = next((row for row in candidates if row.get("safe_to_parse")), None)
    if selected is None and candidates:
        selected = candidates[0]
    payload = _public_payload(candidates, selected)
    payload["report_date"] = report_date
    payload["generated_at"] = _now_iso()
    payload["autopilot"] = True
    payload["manual_file_detected"] = bool(selected and selected.get("safe_to_parse"))
    payload["schema_ohlcv_candidates"] = sum(1 for c in candidates if c.get("schema_ohlcv_candidate"))
    payload["recent_candidates"] = sum(1 for c in candidates if c.get("recent_candidate"))
    lines = [
        "# Manual Data Discovery",
        "",
        f"- candidates_found: {payload['candidates_found']}",
        f"- searched_location_count: {payload['searched_location_count']}",
        f"- safe_to_parse: {str(payload['safe_to_parse']).lower()}",
        f"- xlsx_supported: {str(payload['xlsx_supported']).lower()}",
        f"- contents_printed: {str(payload['contents_printed']).lower()}",
        f"- next_required_action: {payload['next_required_action']}",
        "",
    ]
    if selected:
        lines.extend(
            [
                "## Selected candidate",
                "",
                f"- filename: {selected.get('filename', '-')}",
                f"- extension: {selected.get('extension', '-')}",
                f"- pii_guard_status: {selected.get('pii_guard_status', '-')}",
                "",
            ]
        )
    selected_path = None
    if selected and selected.get("safe_to_parse"):
        selected_path = Path(str(selected["resolved_path"]))
    return ManualDataDiscoveryResult(
        markdown_text="\n".join(lines),
        json_payload=payload,
        selected_path=selected_path,
    )
