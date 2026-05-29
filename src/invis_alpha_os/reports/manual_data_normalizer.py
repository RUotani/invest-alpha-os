"""Normalize manual data files (CSV/TSV/TXT/XLSX) to canonical OHLCV CSV."""

from __future__ import annotations

import csv
import importlib.util
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from invis_alpha_os.reports.manual_csv_normalizer import (
    BROKER_FORMAT_AUTO,
    BROKER_FORMAT_GENERIC,
    build_manual_csv_normalization,
)
from invis_alpha_os.reports.manual_csv_pii_guard import _delimiter_for_path, run_manual_csv_pii_guard
from invis_alpha_os.reports.manual_csv_schema import CANONICAL_COLUMNS
from invis_alpha_os.reports.manual_data_discovery import _openpyxl_available
from invis_alpha_os.reports.manual_file_security import scan_manual_file_security


@dataclass(frozen=True)
class ManualDataNormalizationResult:
    markdown_text: str
    json_payload: dict[str, Any]
    normalized_path: Path | None


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def detect_input_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix == ".tsv":
        return "tsv"
    if suffix == ".txt":
        return "pasted_table"
    if suffix == ".xlsx":
        return "xlsx"
    return "unknown"


def _load_tsv_or_txt_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    errors: list[str] = []
    try:
        text = path.read_text(encoding="utf-8-sig")
    except OSError as exc:
        return [], [f"read_failed:{exc.__class__.__name__}"]
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return [], ["file_empty"]
    delimiter = _delimiter_for_path(path, lines[0])
    reader = csv.DictReader(lines, delimiter=delimiter)
    if reader.fieldnames is None:
        return [], ["missing_header"]
    return [dict(row) for row in reader], errors


def _load_xlsx_rows(path: Path) -> tuple[list[dict[str, str]], list[str], bool]:
    if not _openpyxl_available():
        return [], ["xlsx_not_supported_no_openpyxl"], False
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            workbook.close()
            return [], ["xlsx_empty"], True
        headers = [str(cell).strip() for cell in header_row if cell is not None]
        out: list[dict[str, str]] = []
        for values in rows_iter:
            if values is None:
                continue
            row_map: dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = values[idx] if idx < len(values) else ""
                row_map[header] = "" if value is None else str(value).strip()
            if any(row_map.values()):
                out.append(row_map)
        workbook.close()
        return out, [], True
    except Exception as exc:
        return [], [f"xlsx_read_failed:{exc.__class__.__name__}"], True


def _write_intermediate_csv(rows: list[dict[str, str]], headers: list[str], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return output_path


def build_manual_data_normalization(
    *,
    input_path: Path,
    report_date: str,
    broker_format: str = BROKER_FORMAT_AUTO,
    output_path: Path | None = None,
) -> ManualDataNormalizationResult:
    security = scan_manual_file_security(input_path)
    if security.status != "passed":
        payload = {
            "report_date": report_date,
            "generated_at": _now_iso(),
            "ready_for_validation": False,
            "security_scan": security.json_payload,
            "errors": security.issues,
            "contents_printed": False,
        }
        return ManualDataNormalizationResult(
            markdown_text="# Manual Data Normalization Refused\n\n- security_scan_failed: true\n",
            json_payload=payload,
            normalized_path=None,
        )

    input_type = detect_input_type(input_path)
    xlsx_supported = _openpyxl_available()
    pii = run_manual_csv_pii_guard(input_path) if input_type != "xlsx" else None
    if input_type == "xlsx":
        xlsx_rows, xlsx_errors, _ = _load_xlsx_rows(input_path)
        if xlsx_errors:
            payload = {
                "report_date": report_date,
                "generated_at": _now_iso(),
                "input_type": input_type,
                "parser_used": "none",
                "xlsx_supported": xlsx_supported,
                "ready_for_validation": False,
                "errors": xlsx_errors,
                "contents_printed": False,
            }
            return ManualDataNormalizationResult(
                markdown_text="# Manual Data Normalization Failed\n\n- xlsx_not_supported\n",
                json_payload=payload,
                normalized_path=None,
            )
        from invis_alpha_os.reports.manual_csv_pii_guard import scan_csv_headers_for_pii

        pii_headers = list(xlsx_rows[0].keys()) if xlsx_rows else []
        pii = scan_csv_headers_for_pii(pii_headers)
        work_csv = input_path.parent / f"manual_data_intermediate_{report_date}.csv"
        if not xlsx_rows:
            payload = {
                "report_date": report_date,
                "generated_at": _now_iso(),
                "input_type": input_type,
                "ready_for_validation": False,
                "errors": ["xlsx_no_rows"],
                "contents_printed": False,
            }
            return ManualDataNormalizationResult(
                markdown_text="# Manual Data Normalization Failed\n",
                json_payload=payload,
                normalized_path=None,
            )
        fieldnames = list(xlsx_rows[0].keys())
        _write_intermediate_csv(xlsx_rows, fieldnames, work_csv)
        pipeline_path = work_csv
        parser_used = "openpyxl_to_csv"
        rows = xlsx_rows
    rows: list[dict[str, str]] = []
    if input_type in {"tsv", "pasted_table"}:
        rows, load_errors = _load_tsv_or_txt_rows(input_path)
        if load_errors or not rows:
            payload = {
                "report_date": report_date,
                "generated_at": _now_iso(),
                "input_type": input_type,
                "ready_for_validation": False,
                "errors": load_errors or ["no_rows"],
                "contents_printed": False,
            }
            return ManualDataNormalizationResult(
                markdown_text="# Manual Data Normalization Failed\n",
                json_payload=payload,
                normalized_path=None,
            )
        pii = run_manual_csv_pii_guard(input_path)
        fieldnames = list(rows[0].keys())
        work_csv = input_path.parent / f"manual_data_intermediate_{report_date}.csv"
        _write_intermediate_csv(rows, fieldnames, work_csv)
        pipeline_path = work_csv
        parser_used = "tsv_to_csv" if input_type == "tsv" else "pasted_table_to_csv"
    elif input_type == "csv":
        pipeline_path = input_path
        parser_used = "csv_direct"
        pii = run_manual_csv_pii_guard(input_path)
    else:
        payload = {
            "report_date": report_date,
            "generated_at": _now_iso(),
            "input_type": input_type,
            "ready_for_validation": False,
            "errors": [f"unsupported_input_type:{input_type}"],
            "contents_printed": False,
        }
        return ManualDataNormalizationResult(
            markdown_text="# Manual Data Normalization Refused\n",
            json_payload=payload,
            normalized_path=None,
        )

    if pii and (pii.account_data_detected or pii.status == "rejected"):
        payload = {
            "report_date": report_date,
            "generated_at": _now_iso(),
            "input_type": input_type,
            "pii_guard_status": pii.status,
            "ready_for_validation": False,
            "errors": ["pii_guard_failed"],
            "contents_printed": False,
        }
        return ManualDataNormalizationResult(
            markdown_text="# Manual Data Normalization Refused\n\n- pii_guard_failed: true\n",
            json_payload=payload,
            normalized_path=None,
        )

    norm_out = output_path or pipeline_path.parent / "manual_data_normalized_working.csv"
    broker_for_csv = BROKER_FORMAT_GENERIC if broker_format == BROKER_FORMAT_AUTO else broker_format
    if input_type == "csv":
        csv_result = build_manual_csv_normalization(
            csv_path=pipeline_path,
            report_date=report_date,
            broker_format=broker_for_csv,
            output_path=norm_out,
        )
        payload = dict(csv_result.json_payload)
        payload.update(
            {
                "input_type": input_type,
                "parser_used": parser_used,
                "xlsx_supported": xlsx_supported,
                "contents_printed": False,
            }
        )
        return ManualDataNormalizationResult(
            markdown_text=csv_result.markdown_text.replace("CSV", "Data", 1),
            json_payload=payload,
            normalized_path=csv_result.normalized_path,
        )

    csv_result = build_manual_csv_normalization(
        csv_path=pipeline_path,
        report_date=report_date,
        broker_format=broker_for_csv,
        output_path=norm_out,
    )
    payload = {
        "report_date": report_date,
        "generated_at": _now_iso(),
        "input_type": input_type,
        "parser_used": parser_used,
        "xlsx_supported": xlsx_supported,
        "detected_columns": list(rows[0].keys()) if input_type in {"tsv", "pasted_table"} and rows else [],
        "canonical_columns": list(CANONICAL_COLUMNS),
        "pii_guard_status": pii.status if pii else "unknown",
        "normalized_row_count": csv_result.json_payload.get("normalized_row_count"),
        "ready_for_validation": csv_result.json_payload.get("ready_for_validation"),
        "errors": csv_result.json_payload.get("errors", []),
        "contents_printed": False,
        "cache_write_executed": False,
        "actual_import_executed": False,
    }
    lines = [
        "# Manual Data Normalization",
        "",
        f"- input_type: {input_type}",
        f"- parser_used: {parser_used}",
        f"- ready_for_validation: {str(payload['ready_for_validation']).lower()}",
        "",
    ]
    return ManualDataNormalizationResult(
        markdown_text="\n".join(lines),
        json_payload=payload,
        normalized_path=csv_result.normalized_path,
    )
